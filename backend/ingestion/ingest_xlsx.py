"""
ingest_xlsx.py — Convert ParcelPilot_Assessment_Data.xlsx → SQLite (parcelpilot.db).

Creates three tables:
  accounts  — account metadata
  orders    — order records, each linked to an account_id
  tickets   — support tickets, each linked to an account_id

Column names are normalised (lowercased, spaces → underscores).
"""

import sqlite3
from pathlib import Path

import openpyxl

XLSX_FILE = Path(__file__).parent.parent / "data" / "raw" / "ParcelPilot_Assessment_Data.xlsx"
DB_FILE = Path(__file__).parent.parent / "data" / "parcelpilot.db"


def normalise_col(name: str) -> str:
    """Lower-case, strip, replace spaces/hyphens with underscores."""
    return str(name).strip().lower().replace(" ", "_").replace("-", "_")


def sheet_to_rows(ws) -> tuple[list[str], list[list]]:
    """Return (headers, data_rows) from a worksheet."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [normalise_col(h) if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]
    data = [list(row) for row in rows[1:] if any(v is not None for v in row)]
    return headers, data


def create_table(conn: sqlite3.Connection, table_name: str, headers: list[str], rows: list[list]):
    """Create table and insert all rows."""
    col_defs = ", ".join(f'"{h}" TEXT' for h in headers)
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
    conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
    placeholders = ", ".join("?" for _ in headers)
    conn.executemany(
        f'INSERT INTO "{table_name}" VALUES ({placeholders})',
        [[str(v) if v is not None else None for v in row] for row in rows]
    )
    conn.commit()
    print(f"  [{table_name}] {len(rows)} rows inserted.")


def main():
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    conn = sqlite3.connect(DB_FILE)

    # Map each sheet to a canonical table name
    sheet_map = {
        name: name.strip().lower().replace(" ", "_")
        for name in wb.sheetnames
    }

    for sheet_name, table_name in sheet_map.items():
        ws = wb[sheet_name]
        headers, rows = sheet_to_rows(ws)
        if not headers:
            print(f"  [{sheet_name}] empty, skipping.")
            continue
        create_table(conn, table_name, headers, rows)

    conn.close()
    print(f"\nDatabase saved -> {DB_FILE}")


if __name__ == "__main__":
    main()
